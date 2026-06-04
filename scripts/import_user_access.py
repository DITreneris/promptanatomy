#!/usr/bin/env python3
"""
Bulk import user_access from Promptų Akademijos registration Excel.

Reads name (col B) and email (col C); skips session headers and label rows.
Uses Supabase service role from backend/.env (SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY).

Usage (from repo root):
  pip install openpyxl
  cd backend && pip install -r requirements.txt
  python ../scripts/import_user_access.py "../registracija i Promtu Akademijos mokymus.xlsx" --dry-run
  python ../scripts/import_user_access.py "../registracija i Promtu Akademijos mokymus.xlsx" --apply

Plan values: 3 = modules 1-3 (plan 1, default), 6 = modules 1-6 (plan 2).
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
BACKEND = ROOT / "backend"

# Load env before backend imports
sys.path.insert(0, str(BACKEND))
try:
    from dotenv import load_dotenv

    load_dotenv(BACKEND / ".env")
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

VALID_PLANS = (0, 3, 6, 12, 15)
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
SKIP_B_MARKERS = (
    "vardas pavard",
    "birželio",
    "birzelio",
    "liepos",
    "dalyvavimas",
    "el. pašto",
    "el. pasto",
)


def _normalize_email(raw: str) -> str:
    return raw.strip().lower()


def _is_valid_email(email: str) -> bool:
    return bool(email and EMAIL_RE.match(email))


def _should_skip_name(name: str | None) -> bool:
    if not name or not str(name).strip():
        return True
    lower = str(name).strip().lower()
    return any(m in lower for m in SKIP_B_MARKERS)


def parse_xlsx(path: Path, sheet: str | None = None) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise SystemExit(
            "openpyxl required: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = sheet or wb.sheetnames[0]
    if name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {name!r}. Available: {wb.sheetnames}")
    ws = wb[name]

    rows: list[dict] = []
    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 3:
            continue
        name_val = row[1]
        email_raw = row[2]
        if email_raw is None or not str(email_raw).strip():
            continue
        email = _normalize_email(str(email_raw))
        if not _is_valid_email(email):
            continue
        if _should_skip_name(name_val):
            continue
        rows.append(
            {
                "row": row_num,
                "name": str(name_val).strip() if name_val else "",
                "email": email,
            }
        )
    wb.close()
    return rows


def dedupe_and_warn(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """One row per email; warn when same email maps to different names."""
    by_email: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_email[r["email"]].append(r)

    warnings: list[str] = []
    unique: list[dict] = []
    for email, group in sorted(by_email.items()):
        if len(group) > 1:
            names = ", ".join(f"row {g['row']} ({g['name']!r})" for g in group)
            warnings.append(f"Duplicate email {email}: {names}")
        unique.append(group[-1])
    return unique, warnings


def upsert_participant(email: str, highest_plan: int) -> tuple[str, int | None, int]:
    """Returns (status, previous_plan, new_plan). status: created|updated|unchanged|failed"""
    from db import get_user_access, upsert_user_access

    access = get_user_access(email)
    current = (access["highest_plan"] if access else 0) or 0
    new_plan = max(current, highest_plan)
    if access and current >= new_plan:
        return "unchanged", current, current
    ok = upsert_user_access(email, new_plan, stripe_customer_id=None)
    if not ok:
        return "failed", current, current
    if access is None:
        return "created", None, new_plan
    return "updated", current, new_plan


def main() -> int:
    parser = argparse.ArgumentParser(description="Import user_access from registration Excel")
    parser.add_argument("xlsx", type=Path, help="Path to .xlsx file")
    parser.add_argument(
        "--plan",
        type=int,
        default=3,
        choices=VALID_PLANS,
        help="highest_plan to grant (default 3 = modules 1-3, plan 1)",
    )
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write to Supabase (default: dry-run preview only)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Explicit dry-run (default when --apply is omitted)",
    )
    args = parser.parse_args()
    if args.dry_run and args.apply:
        print("Use either --dry-run or --apply, not both", file=sys.stderr)
        return 1

    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    rows = parse_xlsx(args.xlsx.resolve(), args.sheet)
    unique, warnings = dedupe_and_warn(rows)

    print(f"Parsed {len(rows)} row(s) with email → {len(unique)} unique participant(s)")
    for w in warnings:
        print(f"  WARNING: {w}")

    if args.apply:
        from core.config import get_settings

        if not get_settings().is_supabase_configured():
            print(
                "Supabase not configured. Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in backend/.env",
                file=sys.stderr,
            )
            return 1

    stats = defaultdict(int)
    for p in unique:
        if args.apply:
            status, prev, new = upsert_participant(p["email"], args.plan)
            stats[status] += 1
            extra = ""
            if status == "updated" and prev is not None:
                extra = f" (was {prev} → {new})"
            elif status == "created":
                extra = f" (plan {new})"
            print(f"  [{status}] {p['email']} — {p['name']}{extra}")
        else:
            stats["would_import"] += 1
            print(f"  [dry-run] {p['email']} — {p['name']} → highest_plan={args.plan}")

    print()
    if args.apply:
        print(
            f"Done: created={stats['created']} updated={stats['updated']} "
            f"unchanged={stats['unchanged']} failed={stats['failed']}"
        )
        return 1 if stats["failed"] else 0

    print(f"Dry-run: {stats['would_import']} participant(s). Re-run with --apply to upsert.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
